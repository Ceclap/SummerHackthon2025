import os
import shutil
import logging
import pytesseract
import fitz  # PyMuPDF
from PIL import Image
from typing import Optional, Dict, Any, List
from datetime import datetime
import io
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import base64
import csv
import tempfile

from fastapi import FastAPI, File, UploadFile, Request, Form, HTTPException, Query
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware

from config import config

# Импорт новых модулей
from document_classifier import MoldovanDocumentClassifier, DocumentData
from document_storage import DocumentStorage
from report_generator import ReportGenerator

# Импорт OpenAI
try:
    from openai import OpenAI
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False

# --- Инициализация конфигурации ---
config.create_directories()
config.validate_config()

# Настройка Tesseract
pytesseract.pytesseract.tesseract_cmd = config.get_tesseract_path()

# Настройка логирования
logging.basicConfig(
    level=getattr(logging, config.LOG_LEVEL),
    format=config.LOG_FORMAT,
    handlers=[
        logging.FileHandler(config.LOG_FILE),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Инициализация OpenAI клиента
if OPENAI_AVAILABLE and config.OPENAI_API_KEY:
    openai_client = OpenAI(api_key=config.OPENAI_API_KEY)
else:
    openai_client = None

# Инициализация FastAPI
app = FastAPI(
    title=config.APP_NAME,
    description=config.APP_DESCRIPTION,
    version=config.APP_VERSION
)

# Добавляем CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=config.CORS_ORIGINS,
    allow_credentials=True,
    allow_methods=config.CORS_METHODS,
    allow_headers=config.CORS_HEADERS,
)

# Инициализация новых модулей
document_classifier = MoldovanDocumentClassifier()
document_storage = DocumentStorage()
report_generator = ReportGenerator(document_storage)

templates = Jinja2Templates(directory="templates")

# --- Helper Functions ---

def validate_file(file: UploadFile) -> None:
    """Проверяет валидность загруженного файла."""
    if not file.filename:
        raise HTTPException(status_code=400, detail="Файл не выбран")
    
    file_extension = os.path.splitext(file.filename)[1].lower()
    if file_extension not in config.ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400, 
            detail=f"Неподдерживаемый формат файла. Разрешены: {', '.join(config.ALLOWED_EXTENSIONS)}"
        )

def encode_image_to_base64(image_path: str) -> str:
    """Кодирует изображение в base64 для отправки в OpenAI API."""
    try:
        with open(image_path, "rb") as image_file:
            return base64.b64encode(image_file.read()).decode('utf-8')
    except Exception as e:
        logger.error(f"Ошибка при кодировании изображения: {e}")
        raise HTTPException(status_code=500, detail="Ошибка при обработке изображения")

def extract_text_with_openai(image_path: str) -> str:
    """Извлекает текст из изображения с помощью OpenAI Vision API."""
    if not openai_client:
        raise HTTPException(status_code=500, detail="OpenAI API недоступен")
    
    try:
        logger.info(f"Извлечение текста с OpenAI: {image_path}")
        
        # Кодируем изображение в base64
        base64_image = encode_image_to_base64(image_path)
        
        # Формируем промпт для извлечения данных в CSV формате
        prompt = """
        Извлеки весь текст из этого изображения и представь его в формате CSV.
        
        Инструкции:
        1. Распознай все текстовые данные, включая заголовки, значения, даты, суммы
        2. Структурируй данные в табличном формате
        3. Используй запятые как разделители колонок
        4. Каждая строка должна быть на новой строке
        5. Если есть заголовки таблицы, включи их в первую строку
        6. Сохрани все числовые значения, даты, названия, суммы
        7. Не добавляй лишние символы или форматирование
        
        Верни только CSV данные, без дополнительных комментариев или объяснений.
        """
        
        response = openai_client.chat.completions.create(
            model=config.OPENAI_MODEL,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:image/jpeg;base64,{base64_image}"
                            }
                        }
                    ]
                }
            ],
            max_tokens=config.OPENAI_MAX_TOKENS,
            temperature=config.OPENAI_TEMPERATURE
        )
        
        extracted_text = response.choices[0].message.content.strip()
        logger.info(f"OpenAI извлек текст, длина: {len(extracted_text)}")
        
        return extracted_text
        
    except Exception as e:
        logger.error(f"Ошибка при извлечении текста с OpenAI: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка OpenAI API: {str(e)}")

def extract_text_with_tesseract(image_path: str) -> str:
    """Извлекает текст из изображения с помощью Tesseract OCR (fallback)."""
    try:
        logger.info(f"Извлечение текста с Tesseract: {image_path}")
        image = Image.open(image_path)
        
        if image.mode != 'RGB':
            image = image.convert('RGB')
        
        custom_config = f"{config.TESSERACT_CONFIG} -l {config.TESSERACT_LANGUAGES}"
        text = pytesseract.image_to_string(image, config=custom_config)
        
        logger.info(f"Tesseract извлек текст, длина: {len(text)}")
        return text.strip()
        
    except Exception as e:
        logger.error(f"Ошибка при извлечении текста с Tesseract: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка Tesseract OCR: {str(e)}")

def extract_text_from_image(file_path: str) -> str:
    """Извлекает текст из изображения ТОЛЬКО для конвертации в Excel (использует OpenAI Vision API)."""
    file_extension = os.path.splitext(file_path)[1].lower()
    
    if file_extension in config.IMAGE_EXTENSIONS:
        if openai_client and config.OPENAI_API_KEY:
            try:
                return extract_text_with_openai(file_path)
            except Exception as e:
                logger.error(f"OpenAI Vision API не сработал: {e}")
                raise HTTPException(status_code=500, detail="OpenAI Vision API недоступен или произошла ошибка. Проверьте ваш API ключ и интернет-соединение.")
        else:
            logger.error("OpenAI API ключ не настроен или клиент не инициализирован.")
            raise HTTPException(status_code=500, detail="OpenAI Vision API недоступен. Проверьте ваш API ключ в .env и перезапустите приложение.")
    else:
        raise HTTPException(status_code=400, detail="Файл должен быть изображением")

def extract_text_from_image_for_analysis(file_path: str) -> str:
    """Извлекает текст из изображения для анализа документов (использует Tesseract OCR)."""
    try:
        logger.info(f"Обработка изображения для анализа: {file_path}")
        image = Image.open(file_path)
        
        if image.mode != 'RGB':
            image = image.convert('RGB')
        
        custom_config = f"{config.TESSERACT_CONFIG} -l {config.TESSERACT_LANGUAGES}"
        text = pytesseract.image_to_string(image, config=custom_config)
        
        logger.info(f"Tesseract извлек текст для анализа, длина: {len(text)}")
        return text.strip()
        
    except Exception as e:
        logger.error(f"Ошибка при обработке изображения для анализа: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка при обработке изображения: {str(e)}")

def parse_csv_text_to_data(text: str) -> List[List[str]]:
    """Парсит CSV текст в структурированные данные."""
    try:
        logger.info("Парсинг CSV текста в данные")
        
        # Создаем временный файл для CSV
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, encoding='utf-8') as temp_file:
            temp_file.write(text)
            temp_file_path = temp_file.name
        
        try:
            # Читаем CSV с pandas
            df = pd.read_csv(temp_file_path, header=None, encoding='utf-8')
            data = df.values.tolist()
            
            # Очищаем данные
            cleaned_data = []
            for row in data:
                cleaned_row = [str(cell).strip() if pd.notna(cell) else '' for cell in row]
                if any(cell for cell in cleaned_row):  # Пропускаем пустые строки
                    cleaned_data.append(cleaned_row)
            
            logger.info(f"CSV парсинг завершен. Строк: {len(cleaned_data)}, Колонок: {len(cleaned_data[0]) if cleaned_data else 0}")
            return cleaned_data
            
        finally:
            # Удаляем временный файл
            os.unlink(temp_file_path)
            
    except Exception as e:
        logger.error(f"Ошибка при парсинге CSV: {e}")
        # Fallback: простой парсинг по строкам
        lines = text.strip().split('\n')
        data = []
        for line in lines:
            if line.strip():
                # Разделяем по запятым, но обрабатываем кавычки
                row = []
                current = ''
                in_quotes = False
                for char in line:
                    if char == '"':
                        in_quotes = not in_quotes
                    elif char == ',' and not in_quotes:
                        row.append(current.strip())
                        current = ''
                    else:
                        current += char
                row.append(current.strip())
                data.append(row)
        return data

def convert_image_to_pdf(image_path: str, output_path: str) -> str:
    """Конвертирует изображение в PDF."""
    try:
        logger.info(f"Конвертация изображения в PDF: {image_path} -> {output_path}")
        
        image = Image.open(image_path)
        
        if image.mode != 'RGB':
            image = image.convert('RGB')
        
        image.save(output_path, 'PDF', resolution=100.0)
        
        logger.info(f"Изображение успешно конвертировано в PDF: {output_path}")
        return output_path
        
    except Exception as e:
        logger.error(f"Ошибка при конвертации изображения в PDF: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка при конвертации в PDF: {str(e)}")

def create_excel_file(data: List[List[str]], output_path: str) -> str:
    """Создает Excel файл из структурированных данных."""
    try:
        logger.info(f"Создание Excel файла: {output_path}")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Данные из изображения"
        
        # Добавляем данные
        for row_idx, row_data in enumerate(data, 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)
        
        # Стилизация заголовка (первая строка)
        if data:
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for col in range(1, len(data[0]) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
        
        # Автоматическая ширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Добавляем границы
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=len(data), min_col=1, max_col=len(data[0]) if data else 1):
            for cell in row:
                cell.border = thin_border
        
        wb.save(output_path)
        logger.info(f"Excel файл успешно создан: {output_path}")
        return output_path
        
    except Exception as e:
        logger.error(f"Ошибка при создании Excel файла: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка при создании Excel файла: {str(e)}")

def extract_text_from_pdf(file_path: str) -> str:
    """Извлекает текст из PDF документа."""
    try:
        logger.info(f"Обработка PDF: {file_path}")
        doc = fitz.open(file_path)
        text = ""
        
        for page_num, page in enumerate(doc):
            logger.info(f"Обработка страницы {page_num + 1}")
            page_text = page.get_text()
            text += page_text + "\n"
        
        doc.close()
        logger.info(f"Успешно извлечен текст из PDF, длина: {len(text)}")
        return text.strip()
    except Exception as e:
        logger.error(f"Ошибка при обработке PDF: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка при обработке PDF: {str(e)}")

def classify_document(text: str) -> Dict[str, Any]:
    """
    Классифицирует документ с помощью MoldovanDocumentClassifier.
    Возвращает подробную информацию о типе документа и извлеченных полях.
    """
    try:
        # Используем новый классификатор
        doc_data = document_classifier.process_document(text)
        
        # Валидация документа
        validation_result = document_classifier.validate_document(doc_data)
        
        # Формируем результат в старом формате для совместимости
        result = {
            "type": doc_data.doc_type,
            "confidence": doc_data.confidence * 100,
            "fields": {field.name: field.value for field in doc_data.fields},
            "validation_errors": validation_result.get("errors", []),
            "validation_warnings": validation_result.get("warnings", []),
            "raw_text": doc_data.raw_text
        }
        
        logger.info(f"Документ классифицирован: {doc_data.doc_type} (уверенность: {doc_data.confidence * 100:.1f}%)")
        return result
        
    except Exception as e:
        logger.error(f"Ошибка при классификации документа: {e}")
        return {
            "type": "Не удалось определить тип документа",
            "confidence": 0,
            "fields": {},
            "validation_errors": [f"Ошибка классификации: {str(e)}"],
            "validation_warnings": [],
            "raw_text": text
        }

# --- API Endpoints ---

@app.get("/", response_class=HTMLResponse)
async def read_root(request: Request):
    """Отображает главную страницу с формой загрузки."""
    return templates.TemplateResponse("index.html", {"request": request, "result": None})

@app.post("/upload", response_class=HTMLResponse)
async def upload_file(request: Request, file: UploadFile = File(...)):
    """Обрабатывает загрузку файла, OCR и классификацию."""
    try:
        logger.info(f"Получен файл: {file.filename}")
        
        # Валидация файла
        validate_file(file)
        
        # Проверка размера файла
        file_size = 0
        file_path = os.path.join(config.UPLOADS_DIR, f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
        
        with open(file_path, "wb") as buffer:
            while chunk := file.file.read(8192):
                file_size += len(chunk)
                if file_size > config.MAX_FILE_SIZE:
                    os.remove(file_path)
                    raise HTTPException(status_code=400, detail=f"Файл слишком большой (максимум {config.MAX_FILE_SIZE // (1024*1024)}MB)")
                buffer.write(chunk)
        
        logger.info(f"Файл сохранен: {file_path}, размер: {file_size} байт")
        
        # Извлечение текста
        extracted_text = ""
        file_extension = os.path.splitext(file.filename)[1].lower()

        if file_extension == '.pdf':
            extracted_text = extract_text_from_pdf(file_path)
        elif file_extension in config.IMAGE_EXTENSIONS:
            extracted_text = extract_text_from_image_for_analysis(file_path)
        
        if not extracted_text.strip():
            logger.warning("Текст не был извлечен из документа")
            extracted_text = "Текст не был распознан. Возможно, документ содержит только изображения или имеет низкое качество."

        # Классификация документа
        classification_result = classify_document(extracted_text)
        
        # Удаляем файл после обработки
        os.remove(file_path)
        logger.info("Файл удален после обработки")

        result_data = {
            "doc_type": classification_result["type"],
            "confidence": classification_result["confidence"],
            "text": extracted_text,
            "fields": classification_result["fields"],
            "validation_errors": classification_result["validation_errors"],
            "validation_warnings": classification_result["validation_warnings"],
            "filename": file.filename,
            "file_size": file_size,
            "processing_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }

        logger.info(f"Обработка завершена. Тип: {classification_result['type']}, Уверенность: {classification_result['confidence']}%")
        
        return templates.TemplateResponse("index.html", {"request": request, "result": result_data})
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Неожиданная ошибка: {e}")
        raise HTTPException(status_code=500, detail=f"Внутренняя ошибка сервера: {str(e)}")

@app.post("/convert-to-pdf")
async def convert_image_to_pdf_endpoint(file: UploadFile = File(...)):
    """Конвертирует изображение в PDF."""
    try:
        logger.info(f"Получен файл для конвертации в PDF: {file.filename}")
        
        # Проверяем, что это изображение
        file_extension = os.path.splitext(file.filename)[1].lower()
        
        if file_extension not in config.IMAGE_EXTENSIONS:
            raise HTTPException(
                status_code=400, 
                detail=f"Файл должен быть изображением. Разрешены: {', '.join(config.IMAGE_EXTENSIONS)}"
            )
        
        # Проверка размера файла
        file_size = 0
        image_path = os.path.join(config.UPLOADS_DIR, f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
        
        with open(image_path, "wb") as buffer:
            while chunk := file.file.read(8192):
                file_size += len(chunk)
                if file_size > config.MAX_FILE_SIZE:
                    os.remove(image_path)
                    raise HTTPException(status_code=400, detail=f"Файл слишком большой (максимум {config.MAX_FILE_SIZE // (1024*1024)}MB)")
                buffer.write(chunk)
        
        # Конвертируем в PDF
        pdf_filename = os.path.splitext(file.filename)[0] + ".pdf"
        pdf_path = os.path.join(config.UPLOADS_DIR, f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{pdf_filename}")
        
        convert_image_to_pdf(image_path, pdf_path)
        
        # Удаляем исходное изображение
        os.remove(image_path)
        
        logger.info(f"Конвертация завершена: {pdf_path}")
        
        # Возвращаем PDF файл для скачивания
        return FileResponse(
            path=pdf_path,
            filename=pdf_filename,
            media_type='application/pdf'
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка при конвертации: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка при конвертации: {str(e)}")

@app.post("/convert-to-excel")
async def convert_image_to_excel_endpoint(file: UploadFile = File(...)):
    """Конвертирует изображение в Excel таблицу (только через OpenAI Vision API)."""
    try:
        logger.info(f"Получен файл для конвертации в Excel: {file.filename}")
        
        # Проверяем, что это изображение
        file_extension = os.path.splitext(file.filename)[1].lower()
        
        if file_extension not in config.IMAGE_EXTENSIONS:
            raise HTTPException(
                status_code=400, 
                detail=f"Файл должен быть изображением. Разрешены: {', '.join(config.IMAGE_EXTENSIONS)}"
            )
        
        # Проверка размера файла
        file_size = 0
        image_path = os.path.join(config.UPLOADS_DIR, f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
        
        with open(image_path, "wb") as buffer:
            while chunk := file.file.read(8192):
                file_size += len(chunk)
                if file_size > config.MAX_FILE_SIZE:
                    os.remove(image_path)
                    raise HTTPException(status_code=400, detail=f"Файл слишком большой (максимум {config.MAX_FILE_SIZE // (1024*1024)}MB)")
                buffer.write(chunk)
        
        # Извлекаем текст из изображения только через OpenAI
        text = extract_text_from_image(image_path)
        
        if not text.strip():
            raise HTTPException(status_code=400, detail="Не удалось распознать текст в изображении через OpenAI Vision API")
        
        # Парсим текст в структурированные данные
        data = parse_csv_text_to_data(text)
        
        if not data:
            raise HTTPException(status_code=400, detail="Не удалось структурировать данные из изображения")
        
        # Создаем Excel файл
        excel_filename = os.path.splitext(file.filename)[0] + ".xlsx"
        excel_path = os.path.join(config.UPLOADS_DIR, f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{excel_filename}")
        
        create_excel_file(data, excel_path)
        
        # Удаляем исходное изображение
        os.remove(image_path)
        
        logger.info(f"Конвертация в Excel завершена: {excel_path}")
        
        # Возвращаем Excel файл для скачивания
        return FileResponse(
            path=excel_path,
            filename=excel_filename,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка при конвертации в Excel: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка при конвертации в Excel: {str(e)}")

@app.get("/health")
async def health_check():
    """Проверка состояния сервиса."""
    return {
        "status": "healthy", 
        "timestamp": datetime.now().isoformat(),
        "version": config.APP_VERSION,
        "app_name": config.APP_NAME,
        "openai_available": openai_client is not None,
        "tesseract_available": True
    }

@app.get("/config")
async def get_config():
    """Возвращает текущую конфигурацию приложения (только для разработки)."""
    if not config.DEBUG:
        raise HTTPException(status_code=404, detail="Not found")
    
    return {
        "app_name": config.APP_NAME,
        "version": config.APP_VERSION,
        "max_file_size": config.MAX_FILE_SIZE,
        "allowed_extensions": list(config.ALLOWED_EXTENSIONS),
        "confidence_threshold": config.CONFIDENCE_THRESHOLD,
        "document_types": list(config.DOCUMENT_TYPES.keys()),
        "openai_available": openai_client is not None,
        "openai_model": config.OPENAI_MODEL if openai_client else None
    }

# --- Новые эндпоинты для работы с документами ---

@app.post("/documents/upload")
async def upload_document(file: UploadFile = File(...)):
    """Загружает и обрабатывает документ с сохранением в базу данных."""
    try:
        logger.info(f"Загрузка документа: {file.filename}")
        
        # Валидация файла
        validate_file(file)
        
        # Сохраняем файл
        file_size = 0
        file_path = os.path.join(config.UPLOADS_DIR, f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
        
        with open(file_path, "wb") as buffer:
            while chunk := file.file.read(8192):
                file_size += len(chunk)
                if file_size > config.MAX_FILE_SIZE:
                    os.remove(file_path)
                    raise HTTPException(status_code=400, detail=f"Файл слишком большой")
                buffer.write(chunk)
        
        # Извлекаем текст
        extracted_text = ""
        file_extension = os.path.splitext(file.filename)[1].lower()

        if file_extension == '.pdf':
            extracted_text = extract_text_from_pdf(file_path)
        elif file_extension in config.IMAGE_EXTENSIONS:
            extracted_text = extract_text_from_image_for_analysis(file_path)
        
        if not extracted_text.strip():
            raise HTTPException(status_code=400, detail="Не удалось извлечь текст из документа")
        
        # Обрабатываем документ
        doc_data = document_classifier.process_document(extracted_text)
        validation_result = document_classifier.validate_document(doc_data)
        
        # Сохраняем в базу данных
        doc_id = document_storage.store_document(doc_data, file.filename, file_path, validation_result)
        
        return {
            "id": doc_id,
            "filename": file.filename,
            "doc_type": doc_data.doc_type,
            "confidence": doc_data.confidence,
            "fields": {field.name: field.value for field in doc_data.fields},
            "validation_errors": validation_result.get("errors", []),
            "validation_warnings": validation_result.get("warnings", []),
            "message": "Документ успешно загружен и обработан"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка при загрузке документа: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка при загрузке документа: {str(e)}")

@app.get("/documents")
async def get_documents(
    doc_type: Optional[str] = Query(None, description="Тип документа"),
    idno: Optional[str] = Query(None, description="IDNO для поиска"),
    date_from: Optional[str] = Query(None, description="Дата начала (YYYY-MM-DD)"),
    date_to: Optional[str] = Query(None, description="Дата окончания (YYYY-MM-DD)"),
    amount_min: Optional[float] = Query(None, description="Минимальная сумма"),
    amount_max: Optional[float] = Query(None, description="Максимальная сумма"),
    filename: Optional[str] = Query(None, description="Название файла")
):
    """Поиск документов по различным критериям."""
    try:
        from datetime import date as date_type
        
        # Парсим даты
        parsed_date_from = None
        parsed_date_to = None
        
        if date_from:
            try:
                parsed_date_from = date_type.fromisoformat(date_from)
            except ValueError:
                raise HTTPException(status_code=400, detail="Неверный формат даты начала")
        
        if date_to:
            try:
                parsed_date_to = date_type.fromisoformat(date_to)
            except ValueError:
                raise HTTPException(status_code=400, detail="Неверный формат даты окончания")
        
        # Поиск документов
        documents = document_storage.search_documents(
            doc_type=doc_type,
            idno=idno,
            date_from=parsed_date_from,
            date_to=parsed_date_to,
            amount_min=amount_min,
            amount_max=amount_max,
            filename=filename
        )
        
        # Преобразуем в JSON-совместимый формат
        result = []
        for doc in documents:
            doc_dict = {
                "id": doc.id,
                "filename": doc.filename,
                "doc_type": doc.doc_type,
                "upload_date": doc.upload_date.isoformat(),
                "confidence": doc.confidence,
                "fields": {field.name: field.value for field in doc.fields},
                "validation_errors": doc.validation_errors or [],
                "validation_warnings": doc.validation_warnings or []
            }
            result.append(doc_dict)
        
        return {
            "documents": result,
            "total_count": len(result)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка при поиске документов: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка при поиске документов: {str(e)}")

@app.get("/documents/{doc_id}")
async def get_document(doc_id: int):
    """Получает документ по ID."""
    try:
        document = document_storage.get_document(doc_id)
        
        if not document:
            raise HTTPException(status_code=404, detail="Документ не найден")
        
        return {
            "id": document.id,
            "filename": document.filename,
            "doc_type": document.doc_type,
            "upload_date": document.upload_date.isoformat(),
            "confidence": document.confidence,
            "fields": {field.name: field.value for field in document.fields},
            "validation_errors": document.validation_errors or [],
            "validation_warnings": document.validation_warnings or [],
            "raw_text": document.raw_text
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка при получении документа: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка при получении документа: {str(e)}")

@app.delete("/documents/{doc_id}")
async def delete_document(doc_id: int):
    """Удаляет документ по ID."""
    try:
        success = document_storage.delete_document(doc_id)
        
        if not success:
            raise HTTPException(status_code=404, detail="Документ не найден")
        
        return {"message": "Документ успешно удален"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка при удалении документа: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка при удалении документа: {str(e)}")

# --- Эндпоинты для отчетов ---

@app.get("/reports/summary")
async def get_summary_report(
    month: Optional[int] = Query(None, description="Месяц (1-12)"),
    year: Optional[int] = Query(None, description="Год")
):
    """Генерирует сводный отчет по документам."""
    try:
        if month and (month < 1 or month > 12):
            raise HTTPException(status_code=400, detail="Месяц должен быть от 1 до 12")
        
        if year and (year < 2000 or year > 2100):
            raise HTTPException(status_code=400, detail="Год должен быть от 2000 до 2100")
        
        report_data = report_generator.generate_summary_report(month, year)
        
        return report_data
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка при генерации сводного отчета: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка при генерации отчета: {str(e)}")

@app.get("/reports/detailed")
async def get_detailed_report(
    doc_type: Optional[str] = Query(None, description="Тип документа"),
    date_from: Optional[str] = Query(None, description="Дата начала (YYYY-MM-DD)"),
    date_to: Optional[str] = Query(None, description="Дата окончания (YYYY-MM-DD)")
):
    """Генерирует детальный отчет по документам."""
    try:
        from datetime import date as date_type
        
        # Парсим даты
        parsed_date_from = None
        parsed_date_to = None
        
        if date_from:
            try:
                parsed_date_from = date_type.fromisoformat(date_from)
            except ValueError:
                raise HTTPException(status_code=400, detail="Неверный формат даты начала")
        
        if date_to:
            try:
                parsed_date_to = date_type.fromisoformat(date_to)
            except ValueError:
                raise HTTPException(status_code=400, detail="Неверный формат даты окончания")
        
        report_data = report_generator.generate_detailed_report(
            doc_type=doc_type,
            date_from=parsed_date_from,
            date_to=parsed_date_to
        )
        
        return report_data
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка при генерации детального отчета: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка при генерации отчета: {str(e)}")

@app.get("/reports/fiscal")
async def get_fiscal_report(
    month: int = Query(..., description="Месяц (1-12)"),
    year: int = Query(..., description="Год")
):
    """Генерирует отчет для FISC (налоговой службы Молдовы)."""
    try:
        if month < 1 or month > 12:
            raise HTTPException(status_code=400, detail="Месяц должен быть от 1 до 12")
        
        if year < 2000 or year > 2100:
            raise HTTPException(status_code=400, detail="Год должен быть от 2000 до 2100")
        
        report_data = report_generator.generate_fiscal_report(month, year)
        
        return report_data
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка при генерации отчета FISC: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка при генерации отчета FISC: {str(e)}")

@app.get("/reports/export")
async def export_report(
    report_type: str = Query(..., description="Тип отчета: summary, detailed, fiscal"),
    format: str = Query(..., description="Формат экспорта: excel, csv, json"),
    month: Optional[int] = Query(None, description="Месяц (для summary и fiscal)"),
    year: Optional[int] = Query(None, description="Год (для summary и fiscal)"),
    doc_type: Optional[str] = Query(None, description="Тип документа (для detailed)"),
    date_from: Optional[str] = Query(None, description="Дата начала (для detailed)"),
    date_to: Optional[str] = Query(None, description="Дата окончания (для detailed)")
):
    """Экспортирует отчет в различных форматах."""
    try:
        # Генерируем отчет
        if report_type == "summary":
            report_data = report_generator.generate_summary_report(month, year)
        elif report_type == "detailed":
            from datetime import date as date_type
            parsed_date_from = date_type.fromisoformat(date_from) if date_from else None
            parsed_date_to = date_type.fromisoformat(date_to) if date_to else None
            report_data = report_generator.generate_detailed_report(
                doc_type=doc_type,
                date_from=parsed_date_from,
                date_to=parsed_date_to
            )
        elif report_type == "fiscal":
            report_data = report_generator.generate_fiscal_report(month, year)
        else:
            raise HTTPException(status_code=400, detail="Неизвестный тип отчета")
        
        # Создаем файл для экспорта
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if format == "excel":
            filename = f"report_{report_type}_{timestamp}.xlsx"
            file_path = os.path.join(config.UPLOADS_DIR, filename)
            success = report_generator.export_to_excel(report_data, file_path)
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            
        elif format == "csv":
            filename = f"report_{report_type}_{timestamp}.csv"
            file_path = os.path.join(config.UPLOADS_DIR, filename)
            success = report_generator.export_to_csv(report_data, file_path)
            media_type = "text/csv"
            
        elif format == "json":
            filename = f"report_{report_type}_{timestamp}.json"
            file_path = os.path.join(config.UPLOADS_DIR, filename)
            success = report_generator.export_to_json(report_data, file_path)
            media_type = "application/json"
            
        else:
            raise HTTPException(status_code=400, detail="Неизвестный формат экспорта")
        
        if not success:
            raise HTTPException(status_code=500, detail="Ошибка при экспорте отчета")
        
        return FileResponse(
            path=file_path,
            filename=filename,
            media_type=media_type
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка при экспорте отчета: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка при экспорте отчета: {str(e)}")

# --- Main entry point ---
if __name__ == "__main__":
    import uvicorn
    logger.info(f"Запуск {config.APP_NAME} версии {config.APP_VERSION}")
    logger.info(f"Tesseract путь: {config.get_tesseract_path()}")
    if openai_client:
        logger.info(f"OpenAI модель: {config.OPENAI_MODEL}")
    uvicorn.run(app, host=config.HOST, port=config.PORT)
