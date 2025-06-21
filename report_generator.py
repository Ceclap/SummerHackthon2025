import json
import logging
from typing import Dict, List, Optional, Any
from datetime import datetime, date
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from document_storage import DocumentStorage, StoredDocument
from document_classifier import DocumentField

logger = logging.getLogger(__name__)

class ReportGenerator:
    """Генератор отчетов по документам"""
    
    def __init__(self, storage: DocumentStorage):
        self.storage = storage
    
    def generate_summary_report(self, month: Optional[int] = None, year: Optional[int] = None) -> Dict[str, Any]:
        """Генерирует сводный отчет по документам"""
        try:
            stats = self.storage.get_statistics(month, year)
            
            # Формируем текстовый отчет
            report_text = self._format_summary_text(stats, month, year)
            
            # Подготавливаем данные для экспорта
            export_data = {
                "summary": stats,
                "report_text": report_text,
                "generated_at": datetime.now().isoformat(),
                "period": {
                    "month": month,
                    "year": year
                }
            }
            
            logger.info(f"Сводный отчет сгенерирован для периода: {month}/{year if month and year else 'все время'}")
            return export_data
            
        except Exception as e:
            logger.error(f"Ошибка генерации сводного отчета: {e}")
            return {}
    
    def generate_detailed_report(self, 
                               doc_type: Optional[str] = None,
                               date_from: Optional[date] = None,
                               date_to: Optional[date] = None) -> Dict[str, Any]:
        """Генерирует детальный отчет по документам"""
        try:
            # Получаем документы
            documents = self.storage.search_documents(
                doc_type=doc_type,
                date_from=date_from,
                date_to=date_to
            )
            
            # Структурируем данные
            detailed_data = []
            for doc in documents:
                doc_info = {
                    "id": doc.id,
                    "filename": doc.filename,
                    "doc_type": doc.doc_type,
                    "upload_date": doc.upload_date.isoformat(),
                    "confidence": doc.confidence,
                    "validation_errors": doc.validation_errors or [],
                    "validation_warnings": doc.validation_warnings or [],
                    "fields": {}
                }
                
                # Добавляем поля документа
                for field in doc.fields:
                    doc_info["fields"][field.name] = {
                        "value": field.value,
                        "confidence": field.confidence
                    }
                
                detailed_data.append(doc_info)
            
            report_data = {
                "documents": detailed_data,
                "total_count": len(documents),
                "generated_at": datetime.now().isoformat(),
                "filters": {
                    "doc_type": doc_type,
                    "date_from": date_from.isoformat() if date_from else None,
                    "date_to": date_to.isoformat() if date_to else None
                }
            }
            
            logger.info(f"Детальный отчет сгенерирован: {len(documents)} документов")
            return report_data
            
        except Exception as e:
            logger.error(f"Ошибка генерации детального отчета: {e}")
            return {}
    
    def export_to_excel(self, report_data: Dict[str, Any], output_path: str) -> bool:
        """Экспортирует отчет в Excel"""
        try:
            wb = Workbook()
            
            # Удаляем лист по умолчанию
            wb.remove(wb.active)
            
            # Создаем лист со сводкой
            if "summary" in report_data:
                self._create_summary_sheet(wb, report_data["summary"])
            
            # Создаем лист с деталями
            if "documents" in report_data:
                self._create_details_sheet(wb, report_data["documents"])
            
            # Создаем лист с ошибками
            if "documents" in report_data:
                self._create_errors_sheet(wb, report_data["documents"])
            
            wb.save(output_path)
            logger.info(f"Отчет экспортирован в Excel: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Ошибка экспорта в Excel: {e}")
            return False
    
    def export_to_csv(self, report_data: Dict[str, Any], output_path: str) -> bool:
        """Экспортирует отчет в CSV"""
        try:
            if "documents" not in report_data:
                logger.error("Нет данных документов для экспорта в CSV")
                return False
            
            # Преобразуем данные в DataFrame
            rows = []
            for doc in report_data["documents"]:
                row = {
                    "ID": doc["id"],
                    "Файл": doc["filename"],
                    "Тип документа": doc["doc_type"],
                    "Дата загрузки": doc["upload_date"],
                    "Уверенность": doc["confidence"]
                }
                
                # Добавляем поля документа
                for field_name, field_data in doc["fields"].items():
                    row[f"Поле_{field_name}"] = field_data["value"]
                
                rows.append(row)
            
            df = pd.DataFrame(rows)
            df.to_csv(output_path, index=False, encoding='utf-8-sig')
            
            logger.info(f"Отчет экспортирован в CSV: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Ошибка экспорта в CSV: {e}")
            return False
    
    def export_to_json(self, report_data: Dict[str, Any], output_path: str) -> bool:
        """Экспортирует отчет в JSON"""
        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(report_data, f, ensure_ascii=False, indent=2, default=str)
            
            logger.info(f"Отчет экспортирован в JSON: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Ошибка экспорта в JSON: {e}")
            return False
    
    def generate_fiscal_report(self, month: int, year: int) -> Dict[str, Any]:
        """Генерирует отчет для FISC (налоговой службы Молдовы)"""
        try:
            # Получаем все фактуры за месяц
            documents = self.storage.search_documents(
                doc_type="factura_fiscala",
                date_from=date(year, month, 1),
                date_to=date(year, month + 1, 1) if month < 12 else date(year + 1, 1, 1)
            )
            
            fiscal_data = {
                "period": f"{month:02d}/{year}",
                "total_invoices": len(documents),
                "total_amount": 0.0,
                "total_vat": 0.0,
                "invoices": []
            }
            
            for doc in documents:
                invoice_data = {
                    "number": "",
                    "date": "",
                    "seller_idno": "",
                    "buyer_idno": "",
                    "amount": 0.0,
                    "vat_amount": 0.0
                }
                
                # Извлекаем данные из полей
                for field in doc.fields:
                    if field.name == "number":
                        invoice_data["number"] = field.value
                    elif field.name == "date":
                        invoice_data["date"] = field.value
                    elif field.name == "idno":
                        # Определяем, чей IDNO (продавца или покупателя)
                        if "seller" in field.name or "furnizor" in field.name:
                            invoice_data["seller_idno"] = field.value
                        else:
                            invoice_data["buyer_idno"] = field.value
                    elif field.name == "total_amount":
                        try:
                            invoice_data["amount"] = float(field.value.replace(",", ""))
                            fiscal_data["total_amount"] += invoice_data["amount"]
                        except ValueError:
                            pass
                    elif field.name == "vat_amount":
                        try:
                            invoice_data["vat_amount"] = float(field.value.replace(",", ""))
                            fiscal_data["total_vat"] += invoice_data["vat_amount"]
                        except ValueError:
                            pass
                
                fiscal_data["invoices"].append(invoice_data)
            
            logger.info(f"Отчет FISC сгенерирован: {len(documents)} фактур на сумму {fiscal_data['total_amount']:.2f} MDL")
            return fiscal_data
            
        except Exception as e:
            logger.error(f"Ошибка генерации отчета FISC: {e}")
            return {}
    
    def _format_summary_text(self, stats: Dict[str, Any], month: Optional[int], year: Optional[int]) -> str:
        """Форматирует сводный отчет в текстовом виде"""
        period_text = f"за {month:02d}/{year}" if month and year else "за все время"
        
        text = f"📊 ОТЧЕТ ПО ДОКУМЕНТАМ {period_text.upper()}\n"
        text += "=" * 50 + "\n\n"
        
        text += f"📁 Всего документов: {stats.get('total_documents', 0)}\n\n"
        
        text += "📋 По типам документов:\n"
        for doc_type, count in stats.get('by_type', {}).items():
            amount = stats.get('total_amounts', {}).get(doc_type, 0)
            text += f"  • {self._get_doc_type_name(doc_type)}: {count} шт."
            if amount > 0:
                text += f" (сумма: {amount:.2f} MDL)"
            text += "\n"
        
        if month and year:
            monthly_stats = stats.get('monthly_stats', {})
            if monthly_stats:
                text += f"\n📅 За {month:02d}/{year}:\n"
                for doc_type, count in monthly_stats.items():
                    text += f"  • {self._get_doc_type_name(doc_type)}: {count} шт.\n"
        
        text += f"\n🕒 Отчет сгенерирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}\n"
        
        return text
    
    def _get_doc_type_name(self, doc_type: str) -> str:
        """Возвращает человекочитаемое название типа документа"""
        names = {
            "factura_fiscala": "Factură fiscală",
            "bon_fiscal": "Bon fiscal",
            "stat_plata": "Stat de plată",
            "act_achizitie": "Act de achiziție",
            "ordin_plata": "Ordin de plată",
            "raport_avans": "Raport de avans",
            "unknown": "Неизвестный тип"
        }
        return names.get(doc_type, doc_type)
    
    def _create_summary_sheet(self, wb: Workbook, summary: Dict[str, Any]):
        """Создает лист со сводкой в Excel"""
        ws = wb.create_sheet("Сводка")
        
        # Заголовки
        headers = ["Тип документа", "Количество", "Общая сумма (MDL)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Данные
        row = 2
        for doc_type, count in summary.get('by_type', {}).items():
            amount = summary.get('total_amounts', {}).get(doc_type, 0)
            ws.cell(row=row, column=1, value=self._get_doc_type_name(doc_type))
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=amount)
            row += 1
        
        # Итого
        ws.cell(row=row, column=1, value="ИТОГО")
        ws.cell(row=row, column=2, value=summary.get('total_documents', 0))
        ws.cell(row=row, column=3, value=sum(summary.get('total_amounts', {}).values()))
        
        # Автоподбор ширины столбцов
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_details_sheet(self, wb: Workbook, documents: List[Dict[str, Any]]):
        """Создает лист с деталями документов в Excel"""
        ws = wb.create_sheet("Детали")
        
        if not documents:
            ws.cell(row=1, column=1, value="Нет данных")
            return
        
        # Определяем все возможные поля
        all_fields = set()
        for doc in documents:
            all_fields.update(doc.get('fields', {}).keys())
        
        # Заголовки
        headers = ["ID", "Файл", "Тип документа", "Дата загрузки", "Уверенность"] + list(all_fields)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Данные
        for row, doc in enumerate(documents, 2):
            ws.cell(row=row, column=1, value=doc.get('id'))
            ws.cell(row=row, column=2, value=doc.get('filename'))
            ws.cell(row=row, column=3, value=self._get_doc_type_name(doc.get('doc_type', '')))
            ws.cell(row=row, column=4, value=doc.get('upload_date'))
            ws.cell(row=row, column=5, value=doc.get('confidence'))
            
            # Поля документа
            fields = doc.get('fields', {})
            for col, field_name in enumerate(all_fields, 6):
                field_data = fields.get(field_name, {})
                ws.cell(row=row, column=col, value=field_data.get('value', ''))
        
        # Автоподбор ширины столбцов
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_errors_sheet(self, wb: Workbook, documents: List[Dict[str, Any]]):
        """Создает лист с ошибками валидации в Excel"""
        ws = wb.create_sheet("Ошибки")
        
        # Фильтруем документы с ошибками
        docs_with_errors = [doc for doc in documents if doc.get('validation_errors')]
        
        if not docs_with_errors:
            ws.cell(row=1, column=1, value="Ошибок не найдено")
            return
        
        # Заголовки
        headers = ["ID", "Файл", "Тип документа", "Ошибки", "Предупреждения"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        
        # Данные
        for row, doc in enumerate(docs_with_errors, 2):
            ws.cell(row=row, column=1, value=doc.get('id'))
            ws.cell(row=row, column=2, value=doc.get('filename'))
            ws.cell(row=row, column=3, value=self._get_doc_type_name(doc.get('doc_type', '')))
            ws.cell(row=row, column=4, value="; ".join(doc.get('validation_errors', [])))
            ws.cell(row=row, column=5, value="; ".join(doc.get('validation_warnings', [])))
        
        # Автоподбор ширины столбцов
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width 