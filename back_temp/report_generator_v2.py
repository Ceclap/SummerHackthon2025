"""
Улучшенный генератор отчётов для бухгалтеров Молдовы
Поддерживает различные форматы и стандарты FISC
"""

import os
import json
import csv
import logging
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from config import config
from i18n import i18n
from document_storage import DocumentStorage

logger = logging.getLogger(__name__)

class ReportGeneratorV2:
    """Улучшенный генератор отчётов для Молдовы"""
    
    def __init__(self):
        self.storage = DocumentStorage()
        self.reports_dir = config.REPORTS_DIR
        self.reports_dir.mkdir(parents=True, exist_ok=True)
        
        # Стили для Excel
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate_summary_report(self, start_date: Optional[str] = None, 
                              end_date: Optional[str] = None,
                              language: str = "ru") -> Dict[str, Any]:
        """Генерация сводного отчёта"""
        try:
            # Получение документов
            documents = self.storage.search_documents(
                date_from=start_date,
                date_to=end_date
            )
            
            # Статистика по типам документов
            type_stats = {}
            total_amount = 0.0
            total_vat = 0.0
            
            for doc in documents:
                doc_type = doc.doc_type
                if doc_type not in type_stats:
                    type_stats[doc_type] = {
                        'count': 0,
                        'total_amount': 0.0,
                        'total_vat': 0.0
                    }
                
                type_stats[doc_type]['count'] += 1
                
                # Суммы - извлекаем из полей документа
                amount = 0.0
                vat = 0.0
                for field in doc.fields:
                    if "amount" in field.name.lower():
                        try:
                            amount = float(field.value.replace(",", ""))
                        except (ValueError, AttributeError):
                            pass
                    elif "vat" in field.name.lower():
                        try:
                            vat = float(field.value.replace(",", ""))
                        except (ValueError, AttributeError):
                            pass
                
                type_stats[doc_type]['total_amount'] += amount
                type_stats[doc_type]['total_vat'] += vat
                total_amount += amount
                total_vat += vat
            
            # Формирование отчёта
            report = {
                "report_type": "summary",
                "language": language,
                "generated_date": datetime.now().isoformat(),
                "period": {
                    "start_date": start_date,
                    "end_date": end_date
                },
                "statistics": {
                    "total_documents": len(documents),
                    "total_amount": total_amount,
                    "total_vat": total_vat,
                    "documents_by_type": type_stats
                },
                "summary": {
                    "title": i18n.get_text("summary_report", language),
                    "description": f"Отчёт за период: {start_date or 'начало'} - {end_date or 'конец'}",
                    "currency": "MDL",
                    "currency_symbol": "L"
                }
            }
            
            logger.info(f"Сводный отчёт сгенерирован: {len(documents)} документов")
            return report
            
        except Exception as e:
            logger.error(f"Ошибка генерации сводного отчёта: {e}")
            return {"error": str(e)}
    
    def generate_fiscal_report(self, month: int, year: int, 
                             language: str = "ru") -> Dict[str, Any]:
        """Генерация фискального отчёта для FISC"""
        try:
            # Определение периода
            start_date = datetime(year, month, 1).strftime("%Y-%m-%d")
            if month == 12:
                end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = datetime(year, month + 1, 1) - timedelta(days=1)
            end_date = end_date.strftime("%Y-%m-%d")
            
            # Получение документов
            documents = self.storage.search_documents(
                date_from=start_date,
                date_to=end_date
            )
            
            # Фильтрация фискальных документов
            fiscal_docs = [doc for doc in documents 
                          if doc.doc_type in ['factura_fiscala', 'bon_fiscal']]
            
            # Группировка по компаниям
            companies = {}
            total_sales = 0.0
            total_vat = 0.0
            
            for doc in fiscal_docs:
                # Извлекаем данные из полей документа
                company = "Неизвестная компания"
                amount = 0.0
                vat = 0.0
                idno = ""
                
                for field in doc.fields:
                    if "company" in field.name.lower() or "companie" in field.name.lower():
                        company = field.value
                    elif "amount" in field.name.lower() or "suma" in field.name.lower():
                        try:
                            amount = float(field.value.replace(",", ""))
                        except (ValueError, AttributeError):
                            pass
                    elif "vat" in field.name.lower():
                        try:
                            vat = float(field.value.replace(",", ""))
                        except (ValueError, AttributeError):
                            pass
                    elif "idno" in field.name.lower():
                        idno = field.value
                
                if company not in companies:
                    companies[company] = {
                        'documents': [],
                        'total_amount': 0.0,
                        'total_vat': 0.0,
                        'idno': idno
                    }
                
                companies[company]['documents'].append(doc)
                companies[company]['total_amount'] += amount
                companies[company]['total_vat'] += vat
                total_sales += amount
                total_vat += vat
            
            # Формирование отчёта
            report = {
                "report_type": "fiscal",
                "language": language,
                "generated_date": datetime.now().isoformat(),
                "period": {
                    "month": month,
                    "year": year,
                    "start_date": start_date,
                    "end_date": end_date
                },
                "fiscal_data": {
                    "total_documents": len(fiscal_docs),
                    "total_sales": total_sales,
                    "total_vat": total_vat,
                    "companies": companies
                },
                "fisc_format": {
                    "report_code": "FISC_001",
                    "submission_deadline": self._get_fisc_deadline(month, year),
                    "required_fields": ["company", "idno", "total_amount", "vat_amount"]
                }
            }
            
            logger.info(f"Фискальный отчёт сгенерирован: {len(fiscal_docs)} документов")
            return report
            
        except Exception as e:
            logger.error(f"Ошибка генерации фискального отчёта: {e}")
            return {"error": str(e)}
    
    def generate_detailed_report(self, start_date: Optional[str] = None,
                               end_date: Optional[str] = None,
                               doc_types: Optional[List[str]] = None,
                               language: str = "ru") -> Dict[str, Any]:
        """Генерация детального отчёта"""
        try:
            # Получение документов
            documents = self.storage.search_documents(
                date_from=start_date,
                date_to=end_date
            )
            
            # Фильтрация по типам
            if doc_types:
                documents = [doc for doc in documents 
                           if doc.doc_type in doc_types]
            
            # Детальная информация по каждому документу
            detailed_docs = []
            for doc in documents:
                # Преобразуем поля в словарь для совместимости
                extracted_data = {}
                for field in doc.fields:
                    extracted_data[field.name] = field.value
                
                detailed_doc = {
                    "id": doc.id,
                    "filename": doc.filename,
                    "document_type": doc.doc_type,
                    "processing_date": doc.upload_date.isoformat() if doc.upload_date else None,
                    "confidence": doc.confidence,
                    "extracted_data": extracted_data,
                    "validation_status": not bool(doc.validation_errors),
                    "validation_errors": doc.validation_errors or []
                }
                detailed_docs.append(detailed_doc)
            
            # Статистика
            stats = {
                "total_documents": len(documents),
                "valid_documents": len([doc for doc in documents if not doc.validation_errors]),
                "invalid_documents": len([doc for doc in documents if doc.validation_errors]),
                "average_confidence": sum(doc.confidence for doc in documents) / len(documents) if documents else 0
            }
            
            report = {
                "report_type": "detailed",
                "language": language,
                "generated_date": datetime.now().isoformat(),
                "period": {
                    "start_date": start_date,
                    "end_date": end_date
                },
                "filters": {
                    "document_types": doc_types
                },
                "statistics": stats,
                "documents": detailed_docs
            }
            
            logger.info(f"Детальный отчёт сгенерирован: {len(documents)} документов")
            return report
            
        except Exception as e:
            logger.error(f"Ошибка генерации детального отчёта: {e}")
            return {"error": str(e)}
    
    def generate_custom_report(self, template: str, parameters: Dict[str, Any],
                             language: str = "ru") -> Dict[str, Any]:
        """Генерация пользовательского отчёта"""
        try:
            # Поддерживаемые шаблоны
            templates = {
                "monthly_summary": self._generate_monthly_summary,
                "company_analysis": self._generate_company_analysis,
                "vat_analysis": self._generate_vat_analysis,
                "document_quality": self._generate_quality_report
            }
            
            if template not in templates:
                raise ValueError(f"Неизвестный шаблон: {template}")
            
            report = templates[template](parameters, language)
            logger.info(f"Пользовательский отчёт сгенерирован: {template}")
            return report
            
        except Exception as e:
            logger.error(f"Ошибка генерации пользовательского отчёта: {e}")
            return {"error": str(e)}
    
    def export_report(self, report: Dict[str, Any], format: str = "json",
                     filename: Optional[str] = None) -> str:
        """Экспорт отчёта в различные форматы"""
        try:
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                report_type = report.get("report_type", "report")
                filename = f"{report_type}_{timestamp}.{format}"
            
            file_path = self.reports_dir / filename
            
            if format == "json":
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(report, f, ensure_ascii=False, indent=2)
            
            elif format == "csv":
                self._export_to_csv(report, file_path)
            
            elif format == "excel":
                self._export_to_excel(report, file_path)
            
            elif format == "pdf":
                self._export_to_pdf(report, file_path)
            
            else:
                raise ValueError(f"Неподдерживаемый формат: {format}")
            
            logger.info(f"Отчёт экспортирован: {file_path}")
            return str(file_path)
            
        except Exception as e:
            logger.error(f"Ошибка экспорта отчёта: {e}")
            raise
    
    def _export_to_csv(self, report: Dict[str, Any], file_path: Path):
        """Экспорт в CSV"""
        if report.get("report_type") == "detailed":
            documents = report.get("documents", [])
            if documents:
                fieldnames = list(documents[0].keys())
                with open(file_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(documents)
    
    def _export_to_excel(self, report: Dict[str, Any], file_path: Path):
        """Экспорт в Excel с форматированием"""
        wb = Workbook()
        ws = wb.active
        
        # Заголовок
        title = report.get("summary", {}).get("title", "Отчёт")
        ws.title = title[:31]  # Ограничение Excel
        
        # Добавление данных в зависимости от типа отчёта
        if report.get("report_type") == "summary":
            self._add_summary_to_excel(ws, report)
        elif report.get("report_type") == "fiscal":
            self._add_fiscal_to_excel(ws, report)
        elif report.get("report_type") == "detailed":
            self._add_detailed_to_excel(ws, report)
        
        wb.save(file_path)
    
    def _add_summary_to_excel(self, ws, report):
        """Добавление сводного отчёта в Excel"""
        # Заголовок
        ws['A1'] = report.get("summary", {}).get("title", "Сводный отчёт")
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Статистика
        stats = report.get("statistics", {})
        row = 3
        
        ws[f'A{row}'] = "Всего документов:"
        ws[f'B{row}'] = stats.get("total_documents", 0)
        row += 1
        
        ws[f'A{row}'] = "Общая сумма:"
        ws[f'B{row}'] = f"{stats.get('total_amount', 0):.2f} L"
        row += 1
        
        ws[f'A{row}'] = "Общий НДС:"
        ws[f'B{row}'] = f"{stats.get('total_vat', 0):.2f} L"
        row += 2
        
        # Документы по типам
        ws[f'A{row}'] = "Тип документа"
        ws[f'B{row}'] = "Количество"
        ws[f'C{row}'] = "Сумма"
        ws[f'D{row}'] = "НДС"
        
        for cell in ws[row]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        row += 1
        
        for doc_type, data in stats.get("documents_by_type", {}).items():
            ws[f'A{row}'] = i18n.get_text(doc_type)
            ws[f'B{row}'] = data.get("count", 0)
            ws[f'C{row}'] = f"{data.get('total_amount', 0):.2f} L"
            ws[f'D{row}'] = f"{data.get('total_vat', 0):.2f} L"
            
            for cell in ws[row]:
                cell.border = self.border
            
            row += 1
    
    def _add_fiscal_to_excel(self, ws, report):
        """Добавление фискального отчёта в Excel"""
        # Заголовок
        period = report.get("period", {})
        ws['A1'] = f"Фискальный отчёт за {period.get('month')}/{period.get('year')}"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:E1')
        
        # Компании
        row = 3
        ws[f'A{row}'] = "Компания"
        ws[f'B{row}'] = "IDNO"
        ws[f'C{row}'] = "Количество документов"
        ws[f'D{row}'] = "Общая сумма"
        ws[f'E{row}'] = "Общий НДС"
        
        for cell in ws[row]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        row += 1
        
        for company, data in report.get("fiscal_data", {}).get("companies", {}).items():
            ws[f'A{row}'] = company
            ws[f'B{row}'] = data.get("idno", "")
            ws[f'C{row}'] = len(data.get("documents", []))
            ws[f'D{row}'] = f"{data.get('total_amount', 0):.2f} L"
            ws[f'E{row}'] = f"{data.get('total_vat', 0):.2f} L"
            
            for cell in ws[row]:
                cell.border = self.border
            
            row += 1
    
    def _add_detailed_to_excel(self, ws, report):
        """Добавление детального отчёта в Excel"""
        # Заголовок
        ws['A1'] = "Детальный отчёт документов"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:H1')
        
        # Заголовки колонок
        headers = ["ID", "Файл", "Тип", "Дата обработки", "Уверенность", "Сумма", "НДС", "Статус"]
        row = 3
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        row += 1
        
        # Данные документов
        for doc in report.get("documents", []):
            ws.cell(row=row, column=1, value=doc.get("id"))
            ws.cell(row=row, column=2, value=doc.get("filename"))
            ws.cell(row=row, column=3, value=i18n.get_text(doc.get("document_type")))
            ws.cell(row=row, column=4, value=doc.get("processing_date"))
            ws.cell(row=row, column=5, value=f"{doc.get('confidence', 0):.2f}")
            
            extracted_data = doc.get("extracted_data", {})
            ws.cell(row=row, column=6, value=f"{extracted_data.get('total_amount', 0):.2f} L")
            ws.cell(row=row, column=7, value=f"{extracted_data.get('vat_amount', 0):.2f} L")
            ws.cell(row=row, column=8, value="✓" if doc.get("validation_status") else "✗")
            
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = self.border
            
            row += 1
    
    def _export_to_pdf(self, report: Dict[str, Any], file_path: Path):
        """Экспорт в PDF (заглушка)"""
        # Здесь можно добавить генерацию PDF с помощью библиотеки reportlab
        # Пока сохраняем как JSON
        with open(file_path.with_suffix('.json'), 'w', encoding='utf-8') as f:
            json.dump(report, f, ensure_ascii=False, indent=2)
    
    def _get_fisc_deadline(self, month: int, year: int) -> str:
        """Получение срока подачи отчёта в FISC"""
        # Обычно до 25 числа следующего месяца
        if month == 12:
            deadline = datetime(year + 1, 1, 25)
        else:
            deadline = datetime(year, month + 1, 25)
        return deadline.strftime("%Y-%m-%d")
    
    def _generate_monthly_summary(self, parameters: Dict[str, Any], language: str) -> Dict[str, Any]:
        """Генерация месячного сводного отчёта"""
        month = parameters.get("month", datetime.now().month)
        year = parameters.get("year", datetime.now().year)
        
        start_date = datetime(year, month, 1).strftime("%Y-%m-%d")
        if month == 12:
            end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = datetime(year, month + 1, 1) - timedelta(days=1)
        end_date = end_date.strftime("%Y-%m-%d")
        
        return self.generate_summary_report(start_date, end_date, language)
    
    def _generate_company_analysis(self, parameters: Dict[str, Any], language: str) -> Dict[str, Any]:
        """Генерация анализа по компаниям"""
        company = parameters.get("company")
        start_date = parameters.get("start_date")
        end_date = parameters.get("end_date")
        
        documents = self.storage.search_documents(
            date_from=start_date,
            date_to=end_date
        )
        
        if company:
            documents = [doc for doc in documents 
                        if doc.get('extracted_data', {}).get('company') == company]
        
        # Группировка по компаниям
        companies = {}
        for doc in documents:
            doc_company = doc.get('extracted_data', {}).get('company', 'Неизвестная компания')
            if doc_company not in companies:
                companies[doc_company] = {
                    'documents': [],
                    'total_amount': 0.0,
                    'total_vat': 0.0,
                    'document_types': {}
                }
            
            companies[doc_company]['documents'].append(doc)
            amount = doc.get('extracted_data', {}).get('total_amount', 0)
            vat = doc.get('extracted_data', {}).get('vat_amount', 0)
            companies[doc_company]['total_amount'] += amount
            companies[doc_company]['total_vat'] += vat
            
            doc_type = doc.get('document_type', 'unknown')
            if doc_type not in companies[doc_company]['document_types']:
                companies[doc_company]['document_types'][doc_type] = 0
            companies[doc_company]['document_types'][doc_type] += 1
        
        return {
            "report_type": "company_analysis",
            "language": language,
            "generated_date": datetime.now().isoformat(),
            "period": {"start_date": start_date, "end_date": end_date},
            "companies": companies
        }
    
    def _generate_vat_analysis(self, parameters: Dict[str, Any], language: str) -> Dict[str, Any]:
        """Генерация анализа НДС"""
        start_date = parameters.get("start_date")
        end_date = parameters.get("end_date")
        
        documents = self.storage.search_documents(
            date_from=start_date,
            date_to=end_date
        )
        
        # Анализ НДС
        vat_data = {
            "total_vat": 0.0,
            "vat_by_type": {},
            "vat_by_company": {},
            "vat_rate_analysis": {}
        }
        
        for doc in documents:
            vat_amount = doc.get('extracted_data', {}).get('vat_amount', 0)
            doc_type = doc.get('document_type', 'unknown')
            company = doc.get('extracted_data', {}).get('company', 'Неизвестная компания')
            
            vat_data["total_vat"] += vat_amount
            
            if doc_type not in vat_data["vat_by_type"]:
                vat_data["vat_by_type"][doc_type] = 0.0
            vat_data["vat_by_type"][doc_type] += vat_amount
            
            if company not in vat_data["vat_by_company"]:
                vat_data["vat_by_company"][company] = 0.0
            vat_data["vat_by_company"][company] += vat_amount
        
        return {
            "report_type": "vat_analysis",
            "language": language,
            "generated_date": datetime.now().isoformat(),
            "period": {"start_date": start_date, "end_date": end_date},
            "vat_data": vat_data
        }
    
    def _generate_quality_report(self, parameters: Dict[str, Any], language: str) -> Dict[str, Any]:
        """Генерация отчёта о качестве обработки"""
        start_date = parameters.get("start_date")
        end_date = parameters.get("end_date")
        
        documents = self.storage.search_documents(
            date_from=start_date,
            date_to=end_date
        )
        
        quality_stats = {
            "total_documents": len(documents),
            "valid_documents": len([doc for doc in documents if doc.get('is_valid', False)]),
            "invalid_documents": len([doc for doc in documents if not doc.get('is_valid', False)]),
            "average_confidence": sum(doc.get('type_confidence', 0) for doc in documents) / len(documents) if documents else 0,
            "confidence_distribution": {
                "high": len([doc for doc in documents if doc.get('type_confidence', 0) >= 0.8]),
                "medium": len([doc for doc in documents if 0.5 <= doc.get('type_confidence', 0) < 0.8]),
                "low": len([doc for doc in documents if doc.get('type_confidence', 0) < 0.5])
            },
            "validation_errors": {}
        }
        
        # Анализ ошибок валидации
        for doc in documents:
            errors = doc.get('validation_errors', [])
            for error in errors:
                if error not in quality_stats["validation_errors"]:
                    quality_stats["validation_errors"][error] = 0
                quality_stats["validation_errors"][error] += 1
        
        return {
            "report_type": "quality_report",
            "language": language,
            "generated_date": datetime.now().isoformat(),
            "period": {"start_date": start_date, "end_date": end_date},
            "quality_stats": quality_stats
        }

# Создание глобального экземпляра
report_generator_v2 = ReportGeneratorV2() 